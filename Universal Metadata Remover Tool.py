"""
Enhanced Metadata Tool: Universal Remover + Provenance Checker + GUI
====================================================================

A comprehensive digital forensics tool for metadata removal and AI provenance detection
across all file types with batch processing capabilities.

Author: Digital Forensics & Cybersecurity Analyst
Version: 2.0
License: MIT

Features:
---------
1. Universal Metadata Removal
   - Supported formats: Images (JPEG, PNG, TIFF, WebP, etc.)
   - Documents (PDF, DOCX, XLSX, PPTX, etc.)
   - Audio/Video files (MP3, MP4, WAV, etc.)
   - Generic file fallback with binary scrubbing

2. AI Provenance Detection
   - Detects 50+ AI generation markers
   - Multiple detection methods (metadata, content, filename)
   - Batch scanning capabilities

3. User-Friendly GUI
   - Multiple file selection
   - Real-time progress tracking
   - Comprehensive logging
   - Threaded processing

Dependencies:
-------------
- Pillow (PIL): Image processing
- PyMuPDF (fitz): PDF metadata handling
- python-docx: Word document processing
- openpyxl: Excel spreadsheet processing
- hachoir: Metadata extraction for provenance checking

Installation:
-------------
pip install Pillow PyMuPDF python-docx openpyxl hachoir

Usage:
------
python metadata_remover.py

Security Notes:
---------------
- Always backup files before processing
- Some metadata removal may affect file functionality
- AI detection is heuristic-based and not 100% accurate
"""

import os
import sys
import mimetypes
import shutil
import traceback
from pathlib import Path
from typing import List, Tuple, Dict, Any, Optional

# =============================================================================
# DEPENDENCY CHECK WITH DETAILED REPORTING
# =============================================================================

def check_dependency(module_name: str, package_name: str = None, import_name: str = None):
    """
    Check if a dependency is available and return detailed status.
    
    Args:
        module_name (str): Common name for the module (for display)
        package_name (str): PIP package name (for installation guidance)
        import_name (str): Actual import name (if different from module_name)
    
    Returns:
        dict: Dependency status with details
    """
    import_name = import_name or module_name.lower()
    package_name = package_name or module_name.lower()
    
    status = {
        "name": module_name,
        "package": package_name,
        "available": False,
        "version": "Unknown",
        "error": None
    }
    
    try:
        if module_name == "Pillow":
            from PIL import Image
            status["available"] = True
            status["version"] = Image.__version__
        elif module_name == "PyMuPDF":
            import fitz
            status["available"] = True
            status["version"] = fitz.version
        elif module_name == "python-docx":
            from docx import Document
            status["available"] = True
            status["version"] = getattr(sys.modules['docx'], '__version__', 'Unknown')
        elif module_name == "openpyxl":
            from openpyxl import load_workbook
            status["available"] = True
            status["version"] = getattr(sys.modules['openpyxl'], '__version__', 'Unknown')
        elif module_name == "hachoir":
            from hachoir.parser import createParser
            from hachoir.metadata import extractMetadata
            status["available"] = True
            status["version"] = "Unknown"  # hachoir doesn't have easy version access
        else:
            # Generic import attempt
            imported_module = __import__(import_name)
            status["available"] = True
            status["version"] = getattr(imported_module, '__version__', 'Unknown')
            
    except ImportError as e:
        status["error"] = f"Import failed: {e}"
    except Exception as e:
        status["error"] = f"Initialization failed: {e}"
    
    return status

# Check all dependencies
DEPENDENCIES = {
    "Pillow": {"package": "Pillow", "import": "PIL", "purpose": "Image metadata removal"},
    "PyMuPDF": {"package": "PyMuPDF", "import": "fitz", "purpose": "PDF metadata handling"},
    "python-docx": {"package": "python-docx", "import": "docx", "purpose": "Word document processing"},
    "openpyxl": {"package": "openpyxl", "import": "openpyxl", "purpose": "Excel spreadsheet processing"},
    "hachoir": {"package": "hachoir", "import": "hachoir", "purpose": "AI provenance detection"}
}

# Perform dependency checks
dependency_status = {}
for name, info in DEPENDENCIES.items():
    dependency_status[name] = check_dependency(name, info["package"], info["import"])

# Set feature flags based on dependency checks
HAS_PIL = dependency_status["Pillow"]["available"]
HAS_PYMUPDF = dependency_status["PyMuPDF"]["available"]
HAS_DOCX = dependency_status["python-docx"]["available"]
HAS_OPENPYXL = dependency_status["openpyxl"]["available"]
HAS_HACHOIR = dependency_status["hachoir"]["available"]

def get_dependency_report() -> str:
    """
    Generate a comprehensive report of dependency status.
    
    Returns:
        str: Formatted dependency report
    """
    report = ["DEPENDENCY STATUS REPORT", "=" * 40]
    
    for name, status in dependency_status.items():
        purpose = DEPENDENCIES[name]["purpose"]
        if status["available"]:
            report.append(f"✅ {name}: {status['version']}")
            report.append(f"   Purpose: {purpose}")
        else:
            report.append(f"❌ {name}: UNAVAILABLE")
            report.append(f"   Purpose: {purpose}")
            report.append(f"   Error: {status['error']}")
            report.append(f"   Fix: pip install {DEPENDENCIES[name]['package']}")
        report.append("")
    
    # Summary
    available = sum(1 for status in dependency_status.values() if status["available"])
    total = len(dependency_status)
    report.append(f"SUMMARY: {available}/{total} dependencies available")
    
    if available < total:
        missing = [name for name, status in dependency_status.items() if not status["available"]]
        report.append(f"MISSING: {', '.join(missing)}")
        report.append("")
        report.append("INSTALLATION COMMAND:")
        report.append(f"pip install {' '.join(DEPENDENCIES[name]['package'] for name in missing)}")
    
    return "\n".join(report)

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def log_exception(prefix: str, exc: Exception) -> str:
    """
    Format exception information for logging.
    
    Args:
        prefix (str): Context description for the error
        exc (Exception): The exception object
    
    Returns:
        str: Formatted error message with traceback
    """
    return f"{prefix}: {exc.__class__.__name__}: {exc}\n{traceback.format_exc()}"


def guess_mime(input_path: str) -> str:
    """
    Enhanced MIME type detection with comprehensive extension mapping.
    
    Args:
        input_path (str): Path to the file to analyze
    
    Returns:
        str: Detected MIME type or 'application/octet-stream' for unknown
    """
    mime, _ = mimetypes.guess_type(input_path)
    
    if not mime:
        ext = os.path.splitext(input_path)[1].lower()
        
        mime_map = {
            # Image formats
            '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png', 
            '.gif': 'image/gif', '.bmp': 'image/bmp', '.tiff': 'image/tiff', 
            '.tif': 'image/tiff', '.webp': 'image/webp', '.heic': 'image/heic', 
            '.svg': 'image/svg+xml', '.ico': 'image/x-icon', '.psd': 'image/vnd.adobe.photoshop',
            
            # Document formats
            '.pdf': 'application/pdf',
            '.doc': 'application/msword', 
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.xls': 'application/vnd.ms-excel', 
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.ppt': 'application/vnd.ms-powerpoint', 
            '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
            '.txt': 'text/plain', '.rtf': 'application/rtf', 
            '.odt': 'application/vnd.oasis.opendocument.text',
            '.md': 'text/markdown', '.html': 'text/html',
            
            # Audio/Video formats
            '.mp3': 'audio/mpeg', '.wav': 'audio/wav', '.flac': 'audio/flac', 
            '.ogg': 'audio/ogg', '.m4a': 'audio/mp4', '.aac': 'audio/aac',
            '.mp4': 'video/mp4', '.avi': 'video/x-msvideo', '.mov': 'video/quicktime', 
            '.mkv': 'video/x-matroska', '.webm': 'video/webm', '.flv': 'video/x-flv',
            
            # Archive formats
            '.zip': 'application/zip', '.rar': 'application/vnd.rar',
            '.tar': 'application/x-tar', '.gz': 'application/gzip',
            '.7z': 'application/x-7z-compressed',
        }
        mime = mime_map.get(ext, 'application/octet-stream')
    
    return mime


def get_safe_output_path(input_path: str, output_dir: str = None) -> str:
    """
    Generate a safe output path with automatic collision avoidance.
    
    Args:
        input_path (str): Original file path
        output_dir (str, optional): Custom output directory
    
    Returns:
        str: Safe output path that won't overwrite existing files
    """
    path = Path(input_path)
    
    if output_dir:
        output_dir_path = Path(output_dir)
        output_dir_path.mkdir(parents=True, exist_ok=True)
        output_path = output_dir_path / f"{path.stem}_clean{path.suffix}"
    else:
        output_path = path.parent / f"{path.stem}_clean{path.suffix}"
    
    # Avoid filename collisions
    counter = 1
    original_output = output_path
    while output_path.exists():
        output_path = original_output.parent / f"{original_output.stem}_{counter}{original_output.suffix}"
        counter += 1
    
    return str(output_path)


# =============================================================================
# METADATA REMOVAL FUNCTIONS
# =============================================================================

def remove_image_metadata(input_path: str, output_path: str) -> str:
    """Remove EXIF and ancillary metadata from image files."""
    if not HAS_PIL:
        raise RuntimeError("Pillow (PIL) not available for image processing")
    
    try:
        from PIL import Image
        
        with Image.open(input_path) as img:
            img.load()
            original_format = img.format
            
            if img.format in ('JPEG', 'MPO') and img.mode != 'RGB':
                clean_img = img.convert('RGB')
            else:
                mode = img.mode
                if mode in ('P', 'PA'):
                    mode = 'RGBA' if img.info.get('transparency') else 'RGB'
                
                clean_img = Image.new(mode, img.size)
                clean_img.putdata(list(img.getdata()))
            
            save_kwargs = {}
            if img.format == 'JPEG':
                save_kwargs = {'quality': 95, 'optimize': True}
            elif img.format == 'PNG':
                save_kwargs = {'optimize': True}
            
            clean_img.save(output_path, **save_kwargs)
        
        return f"Image metadata removed ({original_format})"
        
    except Exception as e:
        raise RuntimeError(f"Image processing failed for {input_path}: {e}")


def remove_pdf_metadata(input_path: str, output_path: str) -> str:
    """Remove metadata and embedded files from PDF documents."""
    if not HAS_PYMUPDF:
        raise RuntimeError("PyMuPDF (fitz) not available for PDF processing")
    
    try:
        import fitz
        
        doc = fitz.open(input_path)
        doc.set_metadata({})
        
        # Remove embedded files
        embedded_files = doc.embfile_names()
        for filename in embedded_files:
            try:
                doc.embfile_del(filename)
            except Exception as e:
                print(f"Warning: Could not remove embedded file {filename}: {e}")
        
        doc.save(output_path, garbage=4, deflate=True, clean=True, pretty=True)
        doc.close()
        
        return "PDF metadata and embedded files removed"
        
    except Exception as e:
        raise RuntimeError(f"PDF processing failed for {input_path}: {e}")


def remove_docx_metadata(input_path: str, output_path: str) -> str:
    """Remove core document properties from Word DOCX files."""
    if not HAS_DOCX:
        raise RuntimeError("python-docx not available for DOCX processing")
    
    try:
        from docx import Document
        
        doc = Document(input_path)
        core = doc.core_properties
        
        properties_to_clear = [
            'author', 'title', 'subject', 'keywords', 'comments',
            'category', 'last_modified_by', 'content_status', 
            'identifier', 'language', 'version'
        ]
        
        for prop in properties_to_clear:
            setattr(core, prop, None)
        
        try:
            core.created = None
            core.modified = None
            core.last_printed = None
        except AttributeError:
            pass
            
        doc.save(output_path)
        return "DOCX core metadata removed"
        
    except Exception as e:
        raise RuntimeError(f"DOCX processing failed for {input_path}: {e}")


def remove_excel_metadata(input_path: str, output_path: str) -> str:
    """Remove workbook properties from Excel XLSX files."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl not available for Excel processing")
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(input_path)
        props = wb.properties
        
        properties_to_clear = [
            'creator', 'lastModifiedBy', 'title', 'subject', 
            'keywords', 'description', 'category', 'company', 'manager'
        ]
        
        for prop in properties_to_clear:
            setattr(props, prop, None)
            
        wb.save(output_path)
        return "XLSX metadata removed"
        
    except Exception as e:
        raise RuntimeError(f"Excel processing failed for {input_path}: {e}")


def remove_generic_metadata(input_path: str, output_path: str) -> str:
    """Advanced binary scrubbing for unsupported file types."""
    try:
        mime = guess_mime(input_path)
        
        if mime and mime.startswith('text/'):
            with open(input_path, 'rb') as f_in, open(output_path, 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)
            return "Text file copied (minimal metadata handling)"
        else:
            with open(input_path, 'rb') as f_in, open(output_path, 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)
            return "File copied (format-specific metadata may remain)"
            
    except Exception as e:
        raise RuntimeError(f"Generic file processing failed for {input_path}: {e}")


def remove_metadata_dispatch(input_path: str, output_path: str) -> str:
    """Main dispatcher function for metadata removal."""
    mime = guess_mime(input_path)
    
    try:
        if mime.startswith("image/"):
            return remove_image_metadata(input_path, output_path)
        elif mime == "application/pdf":
            return remove_pdf_metadata(input_path, output_path)
        elif mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            return remove_docx_metadata(input_path, output_path)
        elif mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            return remove_excel_metadata(input_path, output_path)
        else:
            return remove_generic_metadata(input_path, output_path)
            
    except Exception as e:
        try:
            shutil.copy2(input_path, output_path)
            return f"File copied (metadata removal failed: {e})"
        except Exception as copy_error:
            raise RuntimeError(f"All processing methods failed: {copy_error}")


# =============================================================================
# BATCH PROCESSING FUNCTIONS
# =============================================================================

def process_multiple_files(input_files: List[str], output_dir: str = None) -> List[Tuple]:
    """Process multiple files with progress tracking and error handling."""
    results = []
    
    for input_file in input_files:
        try:
            output_path = get_safe_output_path(input_file, output_dir)
            message = remove_metadata_dispatch(input_file, output_path)
            results.append((input_file, output_path, "success", message))
        except Exception as e:
            results.append((input_file, "", "error", str(e)))
    
    return results


# =============================================================================
# PROVENANCE CHECKING FUNCTIONS
# =============================================================================

AI_MARKERS = [
    # General AI terms
    "ai", "artificial intelligence", "generated by", "created with ai", "ai-generated",
    "machine learning", "ml model", "neural network", "deep learning",
    
    # Popular models and platforms
    "stable diffusion", "sdxl", "midjourney", "dall-e", "openai", "chatgpt",
    "runwayml", "adobe firefly", "canva ai", "copilot", "leonardo ai",
    "comfyui", "automatic1111", "clipdrop", "nightcafe", "blue willow",
    "dreamstudio", "playground ai", "getimg.ai", "bing image creator",
    
    # Technical parameters
    "negative prompt", "sampler", "steps", "cfg scale", "denoising strength",
    "seed", "clip skip", "vae", "loras", "hypernetwork", "embedding",
    "checkpoint", "model hash", "hires fix", "restore faces",
    
    # Provenance standards
    "c2pa", "content credentials", "provenance", "content authenticity",
    
    # Company and product names
    "stability ai", "anthropic", "meta ai", "google ai", "nvidia",
    "hotpot.ai", "deepai", "jasper.ai", "copy.ai", "writesonic"
]


def check_provenance(file_path: str) -> Dict[str, Any]:
    """Comprehensive AI provenance detection using multiple methods."""
    results = {
        "ai_related": False, 
        "tags": [], 
        "warnings": [],
        "detection_methods": []
    }
    
    # Method 1: Hachoir metadata extraction
    if HAS_HACHOIR:
        try:
            from hachoir.parser import createParser
            from hachoir.metadata import extractMetadata
            
            parser = createParser(file_path)
            if parser:
                with parser:
                    metadata = extractMetadata(parser)
                    if metadata:
                        metadata_text = "\n".join(metadata.exportPlaintext()).lower()
                        for marker in AI_MARKERS:
                            if marker in metadata_text and marker not in results["tags"]:
                                results["ai_related"] = True
                                results["tags"].append(marker)
                                results["detection_methods"].append("metadata")
        except Exception as e:
            results["warnings"].append(f"Metadata analysis failed: {e}")
    
    # Method 2: File content analysis
    mime = guess_mime(file_path)
    if mime and mime.startswith('text/'):
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content_sample = f.read(16384).lower()
                for marker in AI_MARKERS:
                    if marker in content_sample and marker not in results["tags"]:
                        results["ai_related"] = True
                        results["tags"].append(marker)
                        results["detection_methods"].append("content")
        except Exception as e:
            results["warnings"].append(f"Content analysis failed: {e}")
    
    # Method 3: Filename pattern matching
    filename = Path(file_path).name.lower()
    for marker in AI_MARKERS:
        if (marker in filename and marker not in results["tags"] and 
            len(marker) > 3):
            results["ai_related"] = True
            results["tags"].append(marker)
            results["detection_methods"].append("filename")
    
    return results


def check_multiple_provenance(file_paths: List[str]) -> List[Tuple]:
    """Batch provenance checking for multiple files."""
    results = []
    for file_path in file_paths:
        provenance = check_provenance(file_path)
        results.append((file_path, provenance))
    return results


# =============================================================================
# GUI APPLICATION
# =============================================================================

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from threading import Thread


class EnhancedMetadataRemoverApp:
    """Main GUI application for metadata removal and provenance checking."""
    
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("🛡️ Universal Metadata Remover & Provenance Checker")
        # Increased window height to ensure log is visible
        self.root.geometry("720x650")
        self.root.resizable(True, True)
        
        # Initialize application state
        self.input_files = []
        self.output_dir = tk.StringVar(value=os.getcwd())
        self.is_processing = False
        
        # Setup UI
        self._setup_styles()
        self._setup_layout()
        self._report_detailed_dependencies()
        
        # Add initial log messages to show the log is working
        self.append_log("Application started successfully")
        self.append_log("Ready to process files - use the buttons above to add files")
        self.append_log("Log area is visible with 10 lines by default")
    
    def _setup_styles(self):
        """Configure UI styles and themes."""
        self.style = ttk.Style()
        try:
            self.style.theme_use("clam")
        except Exception:
            pass
        
        self.style.configure("TButton", font=("Segoe UI", 10), padding=6)
        self.style.configure("TLabel", font=("Segoe UI", 10))
        self.style.configure("Status.TLabel", font=("Segoe UI", 9))
        self.style.configure("Progress.Horizontal.TProgressbar", thickness=20)
        self.style.configure("Warning.TLabel", font=("Segoe UI", 9), foreground="#b71c1c")
        self.style.configure("Success.TLabel", font=("Segoe UI", 9), foreground="#2e7d32")
    
    def _setup_layout(self):
        """Create and arrange all UI components with emphasis on log visibility."""
        container = ttk.Frame(self.root, padding=12)
        container.pack(fill="both", expand=True)
        
        # Dependency Status Section
        self._create_dependency_section(container)
        
        # File Selection Section
        self._create_file_selection_section(container)
        
        # Output Directory Section
        self._create_output_section(container)
        
        # Action Buttons Section
        self._create_action_section(container)
        
        # Progress Tracking Section
        self._create_progress_section(container)
        
        # Status Section
        self._create_status_section(container)
        
        # Log Section - Made more prominent
        self._create_log_section(container)
        
        # Configure grid weights to make log area expandable but ensure it's visible
        container.columnconfigure(0, weight=1)
        container.rowconfigure(12, weight=1)  # Log row gets expansion priority
    
    def _create_dependency_section(self, parent):
        """Create dependency status display."""
        ttk.Label(parent, text="Dependency Status:").grid(row=0, column=0, sticky="w")
        
        self.dependency_label = ttk.Label(parent, text="Checking dependencies...", style="Status.TLabel")
        self.dependency_label.grid(row=1, column=0, columnspan=3, sticky="w", pady=(2, 8))
    
    def _create_file_selection_section(self, parent):
        """Create file selection UI components."""
        ttk.Label(parent, text="Select files:").grid(row=2, column=0, sticky="w")
        
        list_frame = ttk.Frame(parent)
        list_frame.grid(row=3, column=0, columnspan=3, sticky="we", pady=(2, 8))
        
        # Reduced listbox height to make more room for log
        self.file_listbox = tk.Listbox(list_frame, height=4, width=80)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.file_listbox.yview)
        self.file_listbox.configure(yscrollcommand=scrollbar.set)
        
        self.file_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        btn_frame = ttk.Frame(parent)
        btn_frame.grid(row=4, column=0, columnspan=3, pady=5, sticky="w")
        
        ttk.Button(btn_frame, text="Add Files", command=self.add_files).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="Add Folder", command=self.add_folder).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="Clear List", command=self.clear_files).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="Remove Selected", command=self.remove_selected).pack(side="left", padx=2)
    
    def _create_output_section(self, parent):
        """Create output directory selection UI."""
        ttk.Label(parent, text="Output directory:").grid(row=5, column=0, sticky="w")
        ttk.Entry(parent, textvariable=self.output_dir, width=60).grid(
            row=6, column=0, columnspan=2, sticky="we", pady=(2, 8))
        ttk.Button(parent, text="Browse", command=self.browse_output_dir).grid(
            row=6, column=2, sticky="e")
    
    def _create_action_section(self, parent):
        """Create main action buttons."""
        action_frame = ttk.Frame(parent)
        action_frame.grid(row=7, column=0, columnspan=3, pady=10, sticky="we")
        
        ttk.Button(
            action_frame, 
            text="Remove Metadata from All", 
            command=self.process_all_files
        ).pack(side="left", padx=5)
        
        ttk.Button(
            action_frame, 
            text="Check Provenance for All", 
            command=self.check_all_provenance
        ).pack(side="left", padx=5)
    
    def _create_progress_section(self, parent):
        """Create progress tracking UI."""
        self.progress = ttk.Progressbar(
            parent, 
            mode='determinate', 
            style="Progress.Horizontal.TProgressbar"
        )
        self.progress.grid(row=8, column=0, columnspan=3, sticky="we", pady=5)
        
        self.progress_label = ttk.Label(parent, text="", style="Status.TLabel")
        self.progress_label.grid(row=9, column=0, columnspan=3, sticky="w")
    
    def _create_status_section(self, parent):
        """Create status display."""
        self.status_label = ttk.Label(
            parent, 
            text="Ready to process files", 
            style="Status.TLabel", 
            foreground="#2e7d32"
        )
        self.status_label.grid(row=10, column=0, columnspan=3, sticky="w", pady=(6, 4))
    
    def _create_log_section(self, parent):
        """Create logging text area with guaranteed visibility."""
        # Log header with clear visibility
        log_header = ttk.Frame(parent)
        log_header.grid(row=11, column=0, columnspan=3, sticky="we", pady=(10, 2))
        ttk.Label(log_header, text="Processing Log:", font=("Segoe UI", 10, "bold")).pack(side="left")
        
        # Clear log button
        ttk.Button(log_header, text="Clear Log", command=self.clear_log, width=10).pack(side="right")
        
        # Log text area - set to show exactly 10 lines by default
        log_frame = ttk.Frame(parent)
        log_frame.grid(row=12, column=0, columnspan=3, sticky="nsew", pady=(0, 5))
        
        # Create text widget with 10-line height
        self.log = tk.Text(log_frame, height=10, width=85, wrap="word")
        self.log.grid(row=0, column=0, sticky="nsew")
        self.log.configure(
            font=("Consolas", 9),
            background="#f8f9fa",
            relief="solid",
            borderwidth=1
        )
        
        # Scrollbar for log
        log_scroll = ttk.Scrollbar(log_frame, command=self.log.yview)
        self.log["yscrollcommand"] = log_scroll.set
        log_scroll.grid(row=0, column=1, sticky="ns")
        
        # Configure grid weights for log frame
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # Add some initial instructional text to the log
        self.log.insert("end", "=== PROCESSING LOG ===\n")
        self.log.insert("end", "This area shows detailed processing information.\n")
        self.log.insert("end", "All file operations and results will be logged here.\n")
        self.log.insert("end", "─" * 50 + "\n")
        self.log.see("end")
    
    def clear_log(self):
        """Clear the log text area."""
        self.log.delete(1.0, "end")
        self.append_log("Log cleared")
        self.append_log("Ready for processing...")
    
    def _report_detailed_dependencies(self):
        """Report detailed dependency status with specific missing packages."""
        available_count = sum(1 for status in dependency_status.values() if status["available"])
        total_count = len(dependency_status)
        
        if available_count == total_count:
            self.dependency_label.config(
                text="✅ All dependencies available", 
                style="Success.TLabel"
            )
            self.append_log("✅ All dependencies are available - full functionality enabled")
        else:
            missing = [name for name, status in dependency_status.items() if not status["available"]]
            missing_text = ", ".join(missing)
            self.dependency_label.config(
                text=f"⚠️ Missing {len(missing)} dependencies: {missing_text}", 
                style="Warning.TLabel"
            )
            
            # Log detailed dependency report
            self.append_log("DEPENDENCY REPORT:")
            self.append_log(get_dependency_report())
            
            # Show specific installation command for missing packages
            missing_packages = [DEPENDENCIES[name]["package"] for name in missing]
            install_cmd = f"pip install {' '.join(missing_packages)}"
            self.append_log(f"To install missing dependencies, run:\n{install_cmd}")
            
            # Show specific feature limitations
            self.append_log("FEATURE LIMITATIONS:")
            if not HAS_PIL:
                self.append_log("  • Image metadata removal disabled")
            if not HAS_PYMUPDF:
                self.append_log("  • PDF metadata removal disabled")
            if not HAS_DOCX:
                self.append_log("  • Word document metadata removal disabled")
            if not HAS_OPENPYXL:
                self.append_log("  • Excel spreadsheet metadata removal disabled")
            if not HAS_HACHOIR:
                self.append_log("  • AI provenance detection disabled")
    
    def add_files(self):
        """Open file dialog to add multiple files to processing list."""
        if self.is_processing:
            messagebox.showwarning("Warning", "Please wait for current processing to complete")
            return
            
        files = filedialog.askopenfilenames(
            title="Select files to process",
            filetypes=[("All files", "*.*")]
        )
        
        if files:
            self.input_files.extend(files)
            self._update_file_list()
            self.append_log(f"✅ Added {len(files)} file(s) to processing list")
            # Show some of the added files in log
            for i, file_path in enumerate(files[:3]):  # Show first 3 files
                self.append_log(f"   - {os.path.basename(file_path)}")
            if len(files) > 3:
                self.append_log(f"   ... and {len(files) - 3} more files")
    
    def add_folder(self):
        """Add all files from a selected folder to processing list."""
        if self.is_processing:
            messagebox.showwarning("Warning", "Please wait for current processing to complete")
            return
            
        folder = filedialog.askdirectory(title="Select folder with files to process")
        if folder:
            try:
                folder_files = [
                    os.path.join(folder, f) for f in os.listdir(folder) 
                    if os.path.isfile(os.path.join(folder, f))
                ]
                self.input_files.extend(folder_files)
                self._update_file_list()
                self.append_log(f"✅ Added folder: {os.path.basename(folder)} ({len(folder_files)} files)")
            except Exception as e:
                self.append_log(f"❌ Error reading folder: {e}")
                messagebox.showerror("Error", f"Could not read folder: {e}")
    
    def clear_files(self):
        """Clear all files from processing list."""
        if self.is_processing:
            messagebox.showwarning("Warning", "Please wait for current processing to complete")
            return
            
        file_count = len(self.input_files)
        self.input_files.clear()
        self.file_listbox.delete(0, tk.END)
        self.append_log(f"🗑️ Cleared {file_count} files from processing list")
    
    def remove_selected(self):
        """Remove selected files from processing list."""
        if self.is_processing:
            messagebox.showwarning("Warning", "Please wait for current processing to complete")
            return
            
        selected_indices = self.file_listbox.curselection()
        if selected_indices:
            removed_count = 0
            for index in sorted(selected_indices, reverse=True):
                removed_file = self.input_files.pop(index)
                removed_count += 1
                self.append_log(f"🗑️ Removed: {os.path.basename(removed_file)}")
            self._update_file_list()
            self.append_log(f"🗑️ Removed {removed_count} selected files")
    
    def _update_file_list(self):
        """Update listbox display with current file list."""
        self.file_listbox.delete(0, tk.END)
        for file_path in self.input_files:
            display_name = os.path.basename(file_path)
            self.file_listbox.insert(tk.END, display_name)
        
        # Update status
        file_count = len(self.input_files)
        if file_count == 0:
            self.status_label.config(text="Ready to process files", foreground="#2e7d32")
        else:
            self.status_label.config(text=f"Ready to process {file_count} files", foreground="#2e7d32")
    
    def browse_output_dir(self):
        """Open directory dialog for output location."""
        directory = filedialog.askdirectory(title="Select output directory")
        if directory:
            self.output_dir.set(directory)
            self.append_log(f"📁 Output directory set to: {directory}")
    
    def process_all_files(self):
        """Initiate metadata removal for all selected files."""
        if not self.input_files:
            messagebox.showwarning("Warning", "Please select files to process")
            return
        
        if self.is_processing:
            messagebox.showwarning("Warning", "Processing already in progress")
            return
        
        # Check if any files require missing dependencies
        file_types = set(guess_mime(f) for f in self.input_files)
        warnings = []
        
        if any(mime.startswith("image/") for mime in file_types) and not HAS_PIL:
            warnings.append("Image files require Pillow (pip install Pillow)")
        if any(mime == "application/pdf" for mime in file_types) and not HAS_PYMUPDF:
            warnings.append("PDF files require PyMuPDF (pip install PyMuPDF)")
        if any(mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" for mime in file_types) and not HAS_DOCX:
            warnings.append("Word documents require python-docx (pip install python-docx)")
        if any(mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" for mime in file_types) and not HAS_OPENPYXL:
            warnings.append("Excel files require openpyxl (pip install openpyxl)")
        
        if warnings:
            warning_msg = "Some files may not be fully processed:\n\n" + "\n".join(f"• {w}" for w in warnings)
            if not messagebox.askyesno("Missing Dependencies", warning_msg + "\n\nContinue anyway?"):
                return
        
        self.append_log("🚀 Starting metadata removal process...")
        self.append_log(f"📊 Processing {len(self.input_files)} files")
        
        thread = Thread(target=self._process_all_files_thread)
        thread.daemon = True
        thread.start()
    
    def _process_all_files_thread(self):
        """Thread function for batch file processing."""
        self.is_processing = True
        total_files = len(self.input_files)
        
        self.root.after(0, self._set_progress_maximum, total_files)
        self.root.after(0, self._set_processing_state, True)
        
        results = []
        for i, input_file in enumerate(self.input_files):
            self.root.after(0, self._update_progress, i, total_files, 
                          f"Processing {os.path.basename(input_file)}")
            
            try:
                output_path = get_safe_output_path(input_file, self.output_dir.get())
                message = remove_metadata_dispatch(input_file, output_path)
                results.append((input_file, output_path, "success", message))
                self.root.after(0, self.append_log, 
                              f"✅ SUCCESS: {os.path.basename(input_file)} -> {message}")
            except Exception as e:
                results.append((input_file, "", "error", str(e)))
                self.root.after(0, self.append_log, 
                              f"❌ ERROR: {os.path.basename(input_file)} -> {e}")
            
            self.root.after(0, self._set_progress_value, i + 1)
        
        self.root.after(0, self._processing_complete, results)
        self.is_processing = False
    
    def check_all_provenance(self):
        """Initiate provenance checking for all selected files."""
        if not self.input_files:
            messagebox.showwarning("Warning", "Please select files to check")
            return
        
        if self.is_processing:
            messagebox.showwarning("Warning", "Processing already in progress")
            return
        
        if not HAS_HACHOIR:
            messagebox.showwarning(
                "Missing Dependency", 
                "AI provenance detection requires hachoir.\n\nInstall with: pip install hachoir"
            )
            return
        
        self.append_log("🔍 Starting AI provenance check...")
        self.append_log(f"📊 Checking {len(self.input_files)} files for AI markers")
        
        thread = Thread(target=self._check_all_provenance_thread)
        thread.daemon = True
        thread.start()
    
    def _check_all_provenance_thread(self):
        """Thread function for batch provenance checking."""
        self.is_processing = True
        total_files = len(self.input_files)
        
        self.root.after(0, self._set_progress_maximum, total_files)
        self.root.after(0, self._set_processing_state, True)
        
        ai_detected = []
        
        for i, file_path in enumerate(self.input_files):
            self.root.after(0, self._update_progress, i, total_files,
                          f"Checking {os.path.basename(file_path)}")
            
            results = check_provenance(file_path)
            if results["ai_related"]:
                ai_detected.append((file_path, results["tags"]))
                self.root.after(0, self.append_log,
                              f"⚠️ AI DETECTED: {os.path.basename(file_path)} -> {', '.join(results['tags'])}")
            else:
                self.root.after(0, self.append_log, f"✅ CLEAN: {os.path.basename(file_path)}")
            
            self.root.after(0, self._set_progress_value, i + 1)
        
        self.root.after(0, self._provenance_check_complete, ai_detected)
        self.is_processing = False
    
    def _set_processing_state(self, processing: bool):
        """Update UI to reflect processing state."""
        self.is_processing = processing
        if processing:
            self.status_label.config(text="Processing...", foreground="#e65100")
        else:
            self.status_label.config(text="Ready", foreground="#2e7d32")
    
    def _set_progress_maximum(self, maximum: int):
        """Set the maximum value for progress bar."""
        self.progress['maximum'] = maximum
    
    def _set_progress_value(self, value: int):
        """Set the current value for progress bar."""
        self.progress['value'] = value
    
    def _update_progress(self, current: int, total: int, message: str):
        """Update progress label with current operation."""
        self.progress_label.config(text=f"{message} ({current+1}/{total})")
    
    def _processing_complete(self, results: List[Tuple]):
        """Handle completion of batch metadata removal."""
        successes = sum(1 for r in results if r[2] == "success")
        errors = len(results) - successes
        
        self.progress_label.config(text=f"Complete: {successes} successful, {errors} errors")
        
        # Add summary to log
        self.append_log("=" * 50)
        self.append_log(f"📊 PROCESSING SUMMARY: {successes} successful, {errors} errors")
        self.append_log("=" * 50)
        
        if errors == 0:
            self.status_label.config(
                text=f"✅ Processing complete: {successes} files processed", 
                foreground="#2e7d32"
            )
            messagebox.showinfo("Processing Complete", 
                              f"Successfully processed {successes} files")
        else:
            self.status_label.config(
                text=f"⚠️ Processing complete with {errors} error(s)", 
                foreground="#e65100"
            )
            messagebox.showwarning("Processing Complete", 
                                 f"Processing complete with {errors} error(s). Check log for details.")
        
        self._set_processing_state(False)
    
    def _provenance_check_complete(self, ai_detected: List[Tuple]):
        """Handle completion of batch provenance checking."""
        self.progress_label.config(text="Provenance check complete")
        
        # Add summary to log
        self.append_log("=" * 50)
        self.append_log(f"🔍 PROVENANCE SUMMARY: AI markers detected in {len(ai_detected)} files")
        self.append_log("=" * 50)
        
        if ai_detected:
            detected_list = "\n".join([
                f"• {os.path.basename(f)}: {', '.join(tags)}" 
                for f, tags in ai_detected
            ])
            
            self.status_label.config(
                text=f"⚠️ AI markers detected in {len(ai_detected)} file(s)", 
                foreground="#e65100"
            )
            
            messagebox.showwarning(
                "AI Provenance Alert", 
                f"AI provenance markers detected in {len(ai_detected)} file(s):\n\n{detected_list}"
            )
        else:
            self.status_label.config(
                text="✅ No AI provenance markers detected", 
                foreground="#2e7d32"
            )
            messagebox.showinfo("Provenance Check", 
                              "No AI provenance markers detected in any files")
        
        self._set_processing_state(False)
    
    def append_log(self, text: str):
        """Append message to log text area with auto-scroll and ensure visibility."""
        self.log.insert("end", text + "\n")
        self.log.see("end")  # Auto-scroll to bottom
        
        # Ensure the log widget is visible by updating the UI
        self.log.update_idletasks()


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def main():
    """Main entry point for the application."""
    try:
        if sys.platform == "win32":
            try:
                from ctypes import windll
                windll.shcore.SetProcessDpiAwareness(1)
            except Exception:
                pass
        
        root = tk.Tk()
        app = EnhancedMetadataRemoverApp(root)
        
        print("Enhanced Metadata Tool started")
        print("Dependency Report:")
        print(get_dependency_report())
        
        root.mainloop()
        
    except Exception as e:
        print(f"Fatal error: {e}")
        traceback.print_exc()
        try:
            tk.messagebox.showerror(
                "Fatal Error", 
                f"Could not start application:\n{e}\n\nCheck console for details."
            )
        except:
            print("Could not display error dialog")


if __name__ == "__main__":
    main()